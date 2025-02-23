import sys
import copy
import os
import pickle

from ncu_report_helper import *

def find_nsight_folder(cuda_base_path):
    for root, dirs, files in os.walk(cuda_base_path):
        if "nsight-compute-" in root and "extras/python" in root:
            return root
    return None

cuda_base_path = "/usr/local/cuda-"
cuda_versions = ["11.0", "11.8", "12.0", "12.1", "12.3"]

ncu_report_path = None

for version in cuda_versions:
    cuda_path = cuda_base_path + version
    result = find_nsight_folder(cuda_path)
    if result:
        ncu_report_path = result
        print(f"\033[0;32;40mUse ncu_report.py moudule in path：{result}.\033[0m")
        break
else:
    print(f"\033[0;31;40mNot found version of cuda that enable ncu_report.py.\033[0m")
    exit()

sys.path.append(ncu_report_path)
import ncu_report

import openpyxl


ncu_report_file_paths = [
    "/home/yangjianchao/Github/nvprof-scripts/ffma_cudacore/NsightCollection/test_ncu_report_file.ncu-rep"
]


if __name__ == "__main__":
    if os.path.exists('ncu_report.xlsx'):
        workbook = openpyxl.load_workbook('ncu_report.xlsx')
    else:
        workbook = openpyxl.Workbook()

    sheet_ncu = workbook.create_sheet('NCU', 1)

    entry = [ \
        "",\
        "Kernel ID", \
        "Thread block Limit SM", \
        "Thread block limit registers", \
        "Thread block limit shared memory", \
        "Thread block limit warps", \
        "theoretical max active warps per SM", \
        "theoretical occupancy", \
        "achieved active warps per SM", \
        "achieved occupancy", \
        "unified L1 cache hit rate", \
        "unified L1 cache hit rate for read transactions (global memory accesses)", \
        "L2 cache hit rate", \
        "GMEM read requests", \
        "GMEM write requests", \
        "GMEM total requests", \
        "GMEM read transactions", \
        "GMEM write transactions", \
        "GMEM total transactions", \
        "number of read transactions per read requests", \
        "number of write transactions per write requests", \
        "L2 read transactions", \
        "L2 write transactions", \
        "L2 total transactions", \
        "DRAM total transactions", \
        "Total number of global atomic requests", \
        "Total number of global reduction requests", \
        "Global memory atomic and reduction transactions", \
        "GPU active cycles", \
        "SM active cycles", \
        "Warp instructions executed", \
        "Instructions executed per clock cycle (IPC)", \
        "Total instructions executed per seconds (MIPS)", \
        "Kernel execution time (ns)", \
    ]
    
    sheet_ncu.append(entry + ["unified L1 cache total requests", "unified L2 cache total requests"])

    for report_file_path in ncu_report_file_paths:
        report = ncu_report.load_report(report_file_path)
        kernel_num = min(len(report[0]), 100)

        for knum in range(kernel_num):
            kernel = report[0][knum]
        
        ################################################################################
        ####                             do ncu report                              ####
        ################################################################################
            print("%81s" % report_file_path.split("/")[-1], "knums: %4d" % \
                len(report[0]), "kernel-%4d" % (knum + 1))
            app_results = [report_file_path.split("/")[-1].split(".")[0], str(knum)]

            app_results.append(get_launch__occupancy_limit_blocks(kernel))
            app_results.append(get_launch__occupancy_limit_registers(kernel))
            app_results.append(get_launch__occupancy_limit_shared_mem(kernel))
            app_results.append(get_launch__occupancy_limit_warps(kernel))
            app_results.append(get_sm__maximum_warps_avg_per_active_cycle(kernel))
            app_results.append(get_sm__maximum_warps_per_active_cycle_pct(kernel))
            app_results.append(get_sm__warps_active_avg_per_cycle_active(kernel))
            app_results.append(get_sm__warps_active_avg_pct_of_peak_sustained_active(kernel))
            app_results.append(get_l1tex__t_sector_hit_rate_pct(kernel))
            app_results.append(get_l1tex__t_sector_pipe_lsu_mem_global_op_ld_hit_rate_pct(kernel))
            app_results.append(get_lts__t_sector_hit_rate_pct(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_ld_sum(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_st_sum(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_ld_sum(kernel) + \
                               get_l1tex__t_requests_pipe_lsu_mem_global_op_st_sum(kernel))
            app_results.append(get_l1tex__t_sectors_pipe_lsu_mem_global_op_ld_sum(kernel))
            app_results.append(get_l1tex__t_sectors_pipe_lsu_mem_global_op_st_sum(kernel))
            app_results.append(get_l1tex__t_sectors_pipe_lsu_mem_global_op_ld_sum(kernel) + \
                               get_l1tex__t_sectors_pipe_lsu_mem_global_op_st_sum(kernel))
            if get_l1tex__t_requests_pipe_lsu_mem_global_op_ld_sum(kernel) == 0.: 
                app_results.append(0.)
            else:
                app_results.append(get_l1tex__t_sectors_pipe_lsu_mem_global_op_ld_sum(kernel) / \
                                   get_l1tex__t_requests_pipe_lsu_mem_global_op_ld_sum(kernel))
            if get_l1tex__t_requests_pipe_lsu_mem_global_op_st_sum(kernel) == 0.: 
                app_results.append(0.)
            else:
                app_results.append(get_l1tex__t_sectors_pipe_lsu_mem_global_op_st_sum(kernel) / \
                                   get_l1tex__t_requests_pipe_lsu_mem_global_op_st_sum(kernel))
            app_results.append(get_l1tex__m_xbar2l1tex_read_sectors_mem_lg_op_ld_sum(kernel))
            app_results.append(get_l1tex__m_l1tex2xbar_write_sectors_mem_lg_op_st_sum(kernel))
            app_results.append(get_l1tex__m_xbar2l1tex_read_sectors_mem_lg_op_ld_sum(kernel) + \
                               get_l1tex__m_l1tex2xbar_write_sectors_mem_lg_op_st_sum(kernel))
            app_results.append(get_dram__sectors_read_sum(kernel) + get_dram__sectors_write_sum(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_atom_sum(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_red_sum(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_atom_sum(kernel) + \
                               get_l1tex__t_requests_pipe_lsu_mem_global_op_red_sum(kernel))
            app_results.append(get_gpc__cycles_elapsed_max(kernel))
            app_results.append(get_sm__cycles_active_avg(kernel))
            app_results.append(get_smsp__inst_executed_sum(kernel))
            app_results.append(get_sm__inst_executed_avg_per_cycle_elapsed(kernel))
            app_results.append(get_sm__inst_executed_avg_per_cycle_elapsed(kernel) * \
                               get_gpc__cycles_elapsed_avg_per_second(kernel) * 1e-6)
            app_results.append(get_gpu__time_duration_sum(kernel))
            
            sheet_ncu.append(app_results + [get_L1_Total_Requests(kernel), \
                                            get_lts__t_requests_srcunit_tex_sum(kernel)])

    # save
    workbook.save('ncu_report.xlsx')
